"""Module for defining various room models."""

from typing import Annotated, Literal, Optional, get_args

from openpyxl import Workbook
from pydantic import Field, RootModel, model_validator
from typing_extensions import Self

from .model_config import BaseConfigModel
from .webpro_options import (
    BuildingType,
    RoomTypeCommunity,
    RoomTypeDepartmentStore,
    RoomTypeFactory,
    RoomTypeHospital,
    RoomTypeHotel,
    RoomTypeOffice,
    RoomTypeResidentialComplex,
    RoomTypeRestaurant,
    RoomTypeSchool,
)


class Zone(BaseConfigModel):
    """Zone info.

    Attributes:
        zone_area: ゾーン面積
    """

    zone_area: Optional[float] = Field(ge=0)


class Zones(RootModel):
    """Zones dict."""

    root: dict[str, Zone]


class Room(BaseConfigModel):
    """全室用途に共通する基底クラス.

    Attributes:
        main_building_type: 建物用途
        building_type: 室用途(大分類)
        room_type: 室用途(小分類)
        floor_height: 階高
        ceiling_height: 天井高
        room_area: 床面積
        zone: ゾーン
        building_model_type: モデル建物
        building_group: 建物群名称
        info: 備考
    """

    main_building_type: Annotated[
        Optional[BuildingType],
        Field(
            alias="mainbuildingType",  # NOTE: for builelib compatibility
        ),
    ] = None

    building_type: Annotated[
        Optional[BuildingType],
        Field(
            alias="buildingType",  # NOTE: for builelib compatibility
        ),
    ] = None

    room_type: Annotated[
        Literal[
            RoomTypeOffice,
            RoomTypeHotel,
            RoomTypeHospital,
            RoomTypeDepartmentStore,
            RoomTypeSchool,
            RoomTypeRestaurant,
            RoomTypeCommunity,
            RoomTypeFactory,
            RoomTypeResidentialComplex,
        ],
        Field(
            alias="roomType",  # NOTE: for builelib compatibility
        ),
    ]

    floor_height: Annotated[
        Optional[float],
        Field(
            gt=0,
            alias="floorHeight",  # NOTE: for builelib compatibility
        ),
    ] = None

    ceiling_height: Annotated[
        Optional[float],
        Field(
            gt=0,
            alias="ceilingHeight",  # NOTE: for builelib compatibility
        ),
    ] = None

    room_area: Annotated[
        Optional[float],
        Field(
            gt=0,
            alias="roomArea",  # NOTE: for builelib compatibility
        ),
    ] = None

    zone: Annotated[
        Optional[Zones],
        Field(
            alias="zone",  # NOTE: for builelib compatibility
        ),
    ] = None

    building_model_type: Annotated[
        Optional[str],
        Field(
            alias="modelBuildingType",  # NOTE: for builelib compatibility
        ),
    ] = None

    building_group: Annotated[
        Optional[str],
        Field(
            alias="buildingGroup",  # NOTE: for builelib compatibility
        ),
    ] = None

    info: Optional[str] = None

    @model_validator(mode="after")
    def _check_building_type_room_type_match(self) -> Self:
        """建物用途と室用途の組み合わせが正しいかチェックする."""
        building_type_to_room_type = {
            "事務所等": RoomTypeOffice,
            "ホテル等": RoomTypeHotel,
            "病院等": RoomTypeHospital,
            "百貨店等": RoomTypeDepartmentStore,
            "学校等": RoomTypeSchool,
            "飲食店等": RoomTypeRestaurant,
            "集会所等": RoomTypeCommunity,
            "工場等": RoomTypeFactory,
            "共同住宅": RoomTypeResidentialComplex,
        }
        valid_room_types = building_type_to_room_type[str(self.building_type)]
        if self.room_type not in get_args(valid_room_types):
            raise ValueError(
                f"room_type '{self.room_type}' is invalid "
                f"for building_type '{self.building_type}'.",
                f"must be one of {valid_room_types} ",
            )
        return self


class Rooms(RootModel):
    """Room dict.

    Attributes:
        root: Room dict.
    """

    root: dict[str, Room]

    @classmethod
    def from_workbook(cls, wb: Workbook, ver: Literal["v2", "v3"] = "v3") -> "Rooms":
        """Create Building instance from workbook.

        Args:
            wb: WEBPRO input workbook instance.
            ver: WEBPRO input workbook version (v2 or v3).
        """
        ws = wb["1) 室仕様"]

        if ver == "v2":
            address = {
                "min_row": 11,
                "floor": "A",
                "name": "B",
                "main_building_type": "C",
                "building_type": "C",
                "room_type": "D",
                "room_area": "E",
                "floor_height": "F",
                "ceiling_height": "G",
                "building_model_type": "L",
                "info": "M",
            }
        elif ver == "v3":
            address = {
                "min_row": 11,
                "floor": "A",
                "name": "B",
                "main_building_type": "C",
                "building_type": "D",
                "room_type": "E",
                "room_area": "F",
                "floor_height": "G",
                "ceiling_height": "H",
                "building_model_type": "M",
                "info": "N",
            }
        else:
            raise ValueError(f"Invalid version: {ver}")

        rooms = {}
        for i_row in range(address["min_row"], 999):
            # break if the first cell is empty or white space
            if (
                ws[f"{address['floor']}{i_row}"].value is None
                or not str(ws[f"{address['floor']}{i_row}"].value).strip()
            ):
                break

            floor = ws[f"{address['floor']}{i_row}"].value
            name = ws[f"{address['name']}{i_row}"].value
            room_key = f"{floor}_{name}"

            rooms[room_key] = Room(
                main_building_type=ws[f"{address['main_building_type']}{i_row}"].value,
                building_type=ws[f"{address['building_type']}{i_row}"].value,
                room_type=ws[f"{address['room_type']}{i_row}"].value,
                room_area=ws[f"{address['room_area']}{i_row}"].value,
                floor_height=ws[f"{address['floor_height']}{i_row}"].value,
                ceiling_height=ws[f"{address['ceiling_height']}{i_row}"].value,
                zone=None,  # TODO: check necessity (builelib)
                building_model_type=ws[f"{address['building_model_type']}{i_row}"].value,
                info=ws[f"{address['info']}{i_row}"].value,
            )

        return cls(root=rooms)
