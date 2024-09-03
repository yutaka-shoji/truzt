"""Module for defining the AirConditioningZone model."""

from typing import Annotated, Literal, Optional, Union

from openpyxl import Workbook
from pydantic import Field, RootModel, field_serializer, field_validator

from .model_config import BaseConfigModel


class AirConditioningZone(BaseConfigModel):
    """Air conditioning zone.

    Attributes:
        is_natual_ventilation: 自然換気の有無
        is_simultaneous_supply: 冷暖同時供給の有無
        ahu_cooling_inside_load: 空調機群名称 冷房室負荷処理
        ahu_cooling_outdoor_load: 空調機群名称 冷房外気負荷処理
        ahu_heating_inside_load: 空調機群名称 暖房室負荷処理
        ahu_heeating_outdoor_load: 空調機群名称 暖房外気負荷処理
        info:
    """

    is_natural_ventilation: Annotated[
        bool,
        Field(
            alias="isNatualVentilation",  # NOTE: builelib タイポ対応
        ),
    ]

    is_simultaneous_supply: Annotated[
        Literal[
            "有",
            "無",
            "有（室負荷）",
            "有（外気負荷）",
        ],
        Field(
            alias="isSimultaneousSupply",  # NOTE: for builelib compatibility
        ),
    ]

    ahu_cooling_inside_load: Annotated[
        str,
        Field(
            alias="AHU_cooling_insideLoad",  # NOTE: for builelib compatibility
        ),
    ]
    ahu_cooling_outdoor_load: Annotated[
        str,
        Field(
            alias="AHU_cooling_outdoorLoad",  # NOTE: for builelib compatibility
        ),
    ]
    ahu_heating_inside_load: Annotated[
        str,
        Field(
            alias="AHU_heating_insideLoad",  # NOTE: for builelib compatibility
        ),
    ]

    ahu_heeating_outdoor_load: Annotated[
        str,
        Field(
            alias="AHU_heating_outdoorLoad",  # NOTE: for builelib compatibility
        ),
    ]

    info: Optional[str]

    @field_validator("is_natural_ventilation", mode="before")
    @classmethod
    def _convert_to_bool(cls, arg: Union[str, bool]) -> bool:
        if arg == "有":
            is_natural_ventilation = True
        elif arg == "無":
            is_natural_ventilation = False
        elif isinstance(arg, bool):
            is_natural_ventilation = arg
        else:
            raise ValueError("is_natural_ventilation must be '有' or '無' or boolean")
        return is_natural_ventilation

    @field_serializer("is_natural_ventilation")
    def _convert_to_str(self, arg: bool) -> Literal["有", "無"]:
        if arg:
            str_arg = "有"
        else:
            str_arg = "無"
        return str_arg


class AirConditioningZones(RootModel):
    """AirConditioningZone dict.

    Attributes:
        root: AirConditioningZone dict.
    """

    root: dict[str, AirConditioningZone]

    @classmethod
    def from_workbook(cls, wb: Workbook, ver: Literal["v2", "v3"] = "v3") -> "AirConditioningZones":
        """Create AirConditioningZones instance from workbook.

        Args:
            wb: WEBPRO input workbook instance.
            ver: WEBPRO input workbook version (v2 or v3).
        """
        ws = wb["2-1) 空調ゾーン"]

        if ver == "v2":
            address = {
                "min_row": 11,
                "floor": "H",
                "name": "I",
                "info": "M",
            }
        elif ver == "v3":
            address = {
                "min_row": 11,
                "floor": "H",
                "name": "I",
                # "is_natural_ventilation": "C",
                # "is_simultaneous_supply": "D",
                "ahu_cooling_inside_load": "J",
                "ahu_cooling_outdoor_load": "K",
                "ahu_heating_inside_load": "J",
                "ahu_heeating_outdoor_load": "K",
                "info": "L",
            }
        else:
            raise ValueError(f"Invalid version: {ver}")

        air_conditioning_zones = {}
        for i_row in range(address["min_row"], 999):
            # break if the first cell is empty or white space
            if (
                ws[f"{address['floor']}{i_row}"].value is None
                or not str(ws[f"{address['floor']}{i_row}"].value).strip()
            ):
                break

            floor = ws[f"{address['floor']}{i_row}"].value
            name = ws[f"{address['name']}{i_row}"].value
            room_key = f"{floor}_{name}"

            air_conditioning_zones[room_key] = AirConditioningZone(
                is_natural_ventilation=False,
                is_simultaneous_supply="無",  # TODO: 暫定
                ahu_cooling_inside_load=ws[f"{address['ahu_cooling_inside_load']}{i_row}"].value,
                ahu_cooling_outdoor_load=ws[f"{address['ahu_cooling_outdoor_load']}{i_row}"].value,
                ahu_heating_inside_load=ws[f"{address['ahu_heating_inside_load']}{i_row}"].value,
                ahu_heeating_outdoor_load=ws[
                    f"{address['ahu_heeating_outdoor_load']}{i_row}"
                ].value,
                info=ws[f"{address['info']}{i_row}"].value,
            )

        return cls(root=air_conditioning_zones)
