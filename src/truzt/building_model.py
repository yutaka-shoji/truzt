"""Module for defining the Building model and related classes."""

from typing import Annotated, Literal, Optional

from openpyxl import Workbook
from pydantic import Field, field_serializer

from .model_config import BaseConfigModel


class BuildingAddress(BaseConfigModel):
    """Building Address.

    Attributes:
        prefecture: 都道府県
        city: 市区町村
        address: 丁目、番地
    """

    prefecture: Optional[str]
    city: Optional[str]
    address: Optional[str]


class CoefficientDHC(BaseConfigModel):
    """Coefficient DHC.

    Attributes:
        cooling: 冷熱
        heating: 温熱
    """

    cooling: Optional[float] = Field(ge=0)
    heating: Optional[float] = Field(ge=0)


class Building(BaseConfigModel):
    """Building info.

    Attributes:
        name: 建築物の名称
        building_address: 建築物所在地
        region: 地域の区分
        annual_solar_region: 年間日射地
        building_floor_area: 延べ面積 (m2)
        coefficient_dhc: 「他人から供給された熱」の一次エネルギー換算係数
    """

    name: str = "Hill Valley High School"

    building_address: Optional[BuildingAddress]

    region: int = Field(ge=1, le=8)

    annual_solar_region: Literal[
        "A1",
        "A2",
        "A3",
        "A4",
        "A5",
    ]

    building_floor_area: float = Field(ge=0)

    coefficient_dhc: Annotated[
        Optional[CoefficientDHC],
        Field(
            alias="Coefficient_DHC",  # for builelib compatibility
        ),
    ] = None

    @field_serializer("region")
    def _serialize_region(self, region: int) -> str:
        return str(region)

    @classmethod
    def from_workbook(cls, wb: Workbook, ver: Literal["v2", "v3"] = "v3") -> "Building":
        """Create Building instance from workbook.

        Args:
            wb: WEBPRO input workbook instance.
            ver: WEBPRO input workbook version (v2 or v3).
        """
        ws = wb["0) 基本情報"]

        if ver == "v2":
            address = {
                "name": "C9",
                "region": "C12",
                "building_address": {
                    "prefecture": "D10",
                    "city": "F10",
                    "address": "C11",
                },
                "building_floor_area": "C17",
                "annual_solar_region": "C18",
                "coefficient_dhc": {
                    "cooling": "C19",
                    "heating": "C20",
                },
            }
        elif ver == "v3":
            address = {
                "name": "C9",
                "region": "C12",
                "building_address": {
                    "prefecture": "D10",
                    "city": "F10",
                    "address": "C11",
                },
                "building_floor_area": "C17",
                "annual_solar_region": "C18",
                "coefficient_dhc": {
                    "cooling": "C19",
                    "heating": "C20",
                },
            }
        else:
            raise ValueError(f"Invalid version: {ver}")

        name = ws[str(address["name"])].value

        building_address = BuildingAddress(
            prefecture=ws[str(address["building_address"]["prefecture"])].value,
            city=ws[str(address["building_address"]["city"])].value,
            address=ws[str(address["building_address"]["address"])].value,
        )

        region = ws[str(address["region"])].value

        building_floor_area = ws[str(address["building_floor_area"])].value

        annual_solar_region = ws[str(address["annual_solar_region"])].value

        coefficient_dhc = CoefficientDHC(
            cooling=ws[str(address["coefficient_dhc"]["cooling"])].value,
            heating=ws[str(address["coefficient_dhc"]["heating"])].value,
        )

        return cls(
            name=name,
            building_address=building_address,
            region=region,
            annual_solar_region=annual_solar_region,
            building_floor_area=building_floor_area,
            coefficient_dhc=coefficient_dhc,
        )
